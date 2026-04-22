"""
remittance_pricer_v3.py  -  FT Partners Remittance Competitive Pricing Tracker

Scrapes live bank-to-bank transfer pricing by intercepting the internal JSON
API calls each provider's website makes when loading their calculator.

Test config: USD->EUR, USD->GBP, USD->MXN at $1,000
Full config:  uncomment CORRIDORS and AMOUNTS at the bottom

REQUIREMENTS:
  pip install playwright openpyxl requests
  playwright install chromium
"""

import re
import time
import json
import argparse
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Default run config (override with --corridors and --amounts flags) ─────────

DEFAULT_CORRIDORS = [
    ("USD", "EUR"),
    ("USD", "GBP"),
    ("USD", "MXN"),
]

DEFAULT_AMOUNTS = [1000]

# Full set — uncomment these to run everything
# DEFAULT_CORRIDORS = [
#     ("USD", "EUR"), ("USD", "GBP"), ("USD", "MXN"),
#     ("USD", "INR"), ("USD", "PHP"), ("USD", "CAD"),
#     ("USD", "AUD"), ("USD", "JPY"),
# ]
# DEFAULT_AMOUNTS = [100, 500, 1000, 5000, 10000]

CURRENCY_TO_COUNTRY = {
    "EUR": "DE", "GBP": "GB", "MXN": "MX", "INR": "IN",
    "PHP": "PH", "CAD": "CA", "AUD": "AU", "JPY": "JP",
}

# Approximate mid-market rates for markup calculation
MID_MARKET_REF = {
    ("USD","EUR"): 0.925, ("USD","GBP"): 0.790, ("USD","MXN"): 17.15,
    ("USD","INR"): 83.50, ("USD","PHP"): 57.50, ("USD","CAD"): 1.365,
    ("USD","AUD"): 1.525, ("USD","JPY"): 149.5,
}

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ── Data model ─────────────────────────────────────────────────────────────────

class Quote:
    def __init__(self, provider, from_ccy, to_ccy, send_amount,
                 fee_usd=None, fx_rate=None, fx_markup_pct=None,
                 received_amount=None, note="", error=""):
        self.provider        = provider
        self.from_ccy        = from_ccy
        self.to_ccy          = to_ccy
        self.send_amount     = send_amount
        self.fee_usd         = fee_usd
        self.fx_rate         = fx_rate
        self.fx_markup_pct   = fx_markup_pct
        self.received_amount = received_amount
        self.note            = note
        self.error           = error

    def fill_gaps(self):
        mid = MID_MARKET_REF.get((self.from_ccy, self.to_ccy))
        if self.fx_markup_pct is None and self.fx_rate and mid:
            self.fx_markup_pct = round((mid - self.fx_rate) / mid * 100, 3)
        if self.received_amount is None and self.fx_rate and self.fee_usd is not None:
            self.received_amount = round((self.send_amount - self.fee_usd) * self.fx_rate, 2)

    def __repr__(self):
        if self.error:
            return f"<Quote {self.provider} ERROR: {self.error}>"
        return (f"<Quote {self.provider} {self.from_ccy}->{self.to_ccy} "
                f"${self.send_amount} fee=${self.fee_usd} "
                f"rate={self.fx_rate} rcv={self.received_amount}>")


# ── Browser setup ──────────────────────────────────────────────────────────────

def make_context(playwright):
    browser = playwright.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-blink-features=AutomationControlled",
        ]
    )
    ctx = browser.new_context(
        user_agent=UA,
        viewport={"width": 1280, "height": 800},
        locale="en-US",
        timezone_id="America/New_York",
    )
    return browser, ctx


# ── Wise ───────────────────────────────────────────────────────────────────────

def scrape_wise(from_ccy, to_ccy, amount, ctx):
    """
    Wise calculator page.
    Intercepts calls to /v3/comparisons or /v1/payment-options which
    contain fee and rate in structured JSON.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        # Target the quotes/payment-options endpoint Wise calls internally
        if any(k in url for k in [
            "/v3/comparisons", "/v2/quotes", "/v1/payment-options",
            "quotes", "comparisons", "price"
        ]):
            try:
                body = response.json()
                results.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    try:
        url = (
            f"https://wise.com/us/send-money/"
            f"?source={from_ccy}&target={to_ccy}&sourceAmount={amount}"
        )
        page.goto(url, wait_until="networkidle", timeout=40000)
        time.sleep(4)
        page.close()

        # Parse captured API responses
        for body in results:
            # Try /v2/quotes response format
            if "paymentOptions" in body:
                options = body["paymentOptions"]
                # Find bank transfer option
                for opt in options:
                    if (opt.get("payIn") == "BANK_TRANSFER" and
                            opt.get("payOut") == "BANK_TRANSFER"):
                        fee = opt.get("fee", {}).get("total", 0)
                        rate = body.get("rate")
                        received = opt.get("targetAmount")
                        if rate:
                            q = Quote(
                                "Wise", from_ccy, to_ccy, amount,
                                fee_usd=float(fee or 0),
                                fx_rate=float(rate),
                                received_amount=float(received) if received else None,
                                note="Wise calculator, bank transfer",
                            )
                            q.fill_gaps()
                            return q

            # Try comparisons format
            if "quotes" in body:
                for q_data in body["quotes"]:
                    pid = (q_data.get("provider", {}).get("id") or "").lower()
                    if "wise" in pid:
                        fee = q_data.get("fee", 0)
                        rate = q_data.get("rate")
                        received = q_data.get("totalReceivedAmount")
                        if rate:
                            q = Quote(
                                "Wise", from_ccy, to_ccy, amount,
                                fee_usd=float(fee or 0),
                                fx_rate=float(rate),
                                received_amount=float(received) if received else None,
                                note="Wise calculator, bank transfer",
                            )
                            q.fill_gaps()
                            return q

        return Quote("Wise", from_ccy, to_ccy, amount,
                     error="Could not parse Wise API response")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("Wise", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── Remitly ────────────────────────────────────────────────────────────────────

def scrape_remitly(from_ccy, to_ccy, amount, ctx):
    """
    Remitly calculator.
    Intercepts calls to their internal pricing/calculator API.
    Economy rate, bank deposit.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        if any(k in url for k in [
            "price", "calculator", "rate", "conduit", "quote", "send-money"
        ]):
            try:
                body = response.json()
                if isinstance(body, dict) and len(body) > 1:
                    results.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    try:
        country = CURRENCY_TO_COUNTRY.get(to_ccy, "DE")
        url = (
            f"https://www.remitly.com/us/en/send-money"
            f"?sourceCountry=USA&destCountry={country}"
            f"&sourceCurrency={from_ccy}&destCurrency={to_ccy}"
            f"&sendAmount={int(amount)}&deliveryType=BANK_DEPOSIT"
        )
        page.goto(url, wait_until="networkidle", timeout=45000)
        time.sleep(5)
        page.close()

        for body in results:
            # Remitly returns a list of rate options
            options = (body.get("rates") or body.get("options") or
                       body.get("conduits") or [])
            if not options and "exchange_rate" in body:
                options = [body]

            for opt in options:
                speed = str(
                    opt.get("speed") or opt.get("delivery_speed") or
                    opt.get("conduit_type") or ""
                ).upper()
                # Prefer Economy (slower, cheaper) over Express
                if "EXPRESS" in speed:
                    continue
                rate = (opt.get("exchange_rate") or opt.get("exchangeRate") or
                        opt.get("rate"))
                fee = (opt.get("fee") or opt.get("transfer_fee") or
                       opt.get("transferFee") or 0)
                received = (opt.get("destination_amount") or
                            opt.get("destinationAmount") or
                            opt.get("recipient_amount") or
                            opt.get("recipientAmount"))
                if rate:
                    q = Quote(
                        "Remitly", from_ccy, to_ccy, amount,
                        fee_usd=float(fee or 0),
                        fx_rate=float(rate),
                        received_amount=float(received) if received else None,
                        note="Economy, bank deposit",
                    )
                    q.fill_gaps()
                    return q

            # If no Economy found, take any rate
            for opt in options:
                rate = (opt.get("exchange_rate") or opt.get("exchangeRate") or
                        opt.get("rate"))
                fee = opt.get("fee") or opt.get("transfer_fee") or 0
                received = (opt.get("destination_amount") or
                            opt.get("destinationAmount") or
                            opt.get("recipient_amount"))
                if rate:
                    q = Quote(
                        "Remitly", from_ccy, to_ccy, amount,
                        fee_usd=float(fee or 0),
                        fx_rate=float(rate),
                        received_amount=float(received) if received else None,
                        note="Bank deposit",
                    )
                    q.fill_gaps()
                    return q

        return Quote("Remitly", from_ccy, to_ccy, amount,
                     error="Could not parse Remitly pricing response")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("Remitly", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── Western Union ──────────────────────────────────────────────────────────────

def scrape_western_union(from_ccy, to_ccy, amount, ctx):
    """
    Western Union price estimator.
    Intercepts their internal pricing API call.
    Bank account pay-in and pay-out.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        if any(k in url for k in [
            "price", "estimate", "fee", "rate", "transfer",
            "pricinginfo", "send-money", "wuenvironments"
        ]):
            try:
                body = response.json()
                if isinstance(body, dict) and len(body) > 1:
                    results.append((url, body))
            except Exception:
                pass

    page.on("response", on_response)

    try:
        country = CURRENCY_TO_COUNTRY.get(to_ccy, "DE")
        url = (
            f"https://www.westernunion.com/us/en/send-money/app/price-estimator"
            f"?sendCurrencyCode={from_ccy}&destCurrencyCode={to_ccy}"
            f"&destCountryCode={country}&sendAmount={int(amount)}"
            f"&payinMethod=ACCOUNT&payoutMethod=ACCOUNT"
        )
        page.goto(url, wait_until="networkidle", timeout=50000)
        time.sleep(5)
        page.close()

        for url, body in results:
            # WU wraps results in paymentOptions or similar
            options = (body.get("paymentOptions") or body.get("pricingOptions") or
                       body.get("options") or [body])
            for opt in options:
                # Look for bank account payout option
                payout = str(opt.get("payoutMethod") or opt.get("payout") or "").upper()
                if payout and "ACCOUNT" not in payout and "BANK" not in payout:
                    continue
                rate = (opt.get("fxRate") or opt.get("exchangeRate") or
                        opt.get("rate") or body.get("fxRate"))
                fee = (opt.get("fee") or opt.get("transferFee") or
                       opt.get("transactionFee") or
                       body.get("fee") or 0)
                received = (opt.get("destPrincipalAmount") or
                            opt.get("receiveAmount") or
                            opt.get("recipientAmount") or
                            body.get("destPrincipalAmount"))
                if rate:
                    q = Quote(
                        "Western Union", from_ccy, to_ccy, amount,
                        fee_usd=float(fee or 0),
                        fx_rate=float(rate),
                        received_amount=float(received) if received else None,
                        note="Bank account delivery (online)",
                    )
                    q.fill_gaps()
                    return q

        return Quote("Western Union", from_ccy, to_ccy, amount,
                     error="Could not parse WU pricing response")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("Western Union", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── MoneyGram ──────────────────────────────────────────────────────────────────

def scrape_moneygram(from_ccy, to_ccy, amount, ctx):
    """
    MoneyGram fee estimator.
    Intercepts their internal fee/rate API.
    Bank deposit delivery.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        if any(k in url for k in [
            "fee", "estimate", "rate", "price", "quote",
            "moneygram", "mgo", "transfer"
        ]):
            try:
                body = response.json()
                if isinstance(body, dict) and len(body) > 1:
                    results.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    try:
        country = CURRENCY_TO_COUNTRY.get(to_ccy, "DE")
        url = (
            f"https://www.moneygram.com/mgo/us/en/fee-estimator"
            f"?sendCurrency={from_ccy}&receiveCurrency={to_ccy}"
            f"&receiveCountry={country}&sendAmount={int(amount)}"
            f"&paymentMethod=BANK_ACCOUNT"
            f"&deliveryMethod=RECEIVE_MONEY_IN_BANK_ACCOUNT"
        )
        page.goto(url, wait_until="networkidle", timeout=50000)
        time.sleep(5)
        page.close()

        for body in results:
            rate = (body.get("exchangeRate") or body.get("fxRate") or
                    body.get("rate") or body.get("receiveExchangeRate"))
            fee = (body.get("fee") or body.get("transferFee") or
                   body.get("mgiSendFee") or body.get("totalFee") or 0)
            received = (body.get("receiveAmount") or body.get("destinationAmount") or
                        body.get("estimatedReceiveAmount") or
                        body.get("receiverAmount"))
            if rate:
                q = Quote(
                    "MoneyGram", from_ccy, to_ccy, amount,
                    fee_usd=float(fee or 0),
                    fx_rate=float(rate),
                    received_amount=float(received) if received else None,
                    note="Bank deposit delivery",
                )
                q.fill_gaps()
                return q

        return Quote("MoneyGram", from_ccy, to_ccy, amount,
                     error="Could not parse MoneyGram pricing response")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("MoneyGram", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── Revolut ────────────────────────────────────────────────────────────────────

def scrape_revolut(from_ccy, to_ccy, amount, ctx):
    """
    Revolut transfer page.
    Standard (free) plan, external bank transfer.
    Intercepts their internal rate/fee API.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        if any(k in url for k in [
            "transfer", "quote", "rate", "price", "fee",
            "exchange", "revolut"
        ]):
            try:
                body = response.json()
                if isinstance(body, dict) and len(body) > 1:
                    results.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    try:
        url = (
            f"https://www.revolut.com/en-US/money-transfer/"
            f"{from_ccy.lower()}-to-{to_ccy.lower()}/"
        )
        page.goto(url, wait_until="networkidle", timeout=40000)
        time.sleep(4)
        page.close()

        for body in results:
            rate = (body.get("rate") or body.get("exchangeRate") or
                    body.get("fxRate") or body.get("midRate"))
            fee = (body.get("fee") or body.get("transferFee") or
                   body.get("totalFee"))
            received = (body.get("recipientAmount") or body.get("receiveAmount") or
                        body.get("targetAmount") or body.get("amount"))
            if rate:
                # Revolut standard plan charges ~$5 for external bank transfers
                # if fee not found in API response
                if fee is None:
                    fee = 5.0
                q = Quote(
                    "Revolut", from_ccy, to_ccy, amount,
                    fee_usd=float(fee),
                    fx_rate=float(rate),
                    received_amount=float(received) if received else None,
                    note="Standard plan, external bank transfer",
                )
                q.fill_gaps()
                return q

        return Quote("Revolut", from_ccy, to_ccy, amount,
                     error="Could not parse Revolut pricing response")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("Revolut", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── Euronet ────────────────────────────────────────────────────────────────────

def scrape_euronet(from_ccy, to_ccy, amount, ctx):
    """
    Euronet money transfer page.
    Intercepts their rate API.
    """
    page = ctx.new_page()
    results = []

    def on_response(response):
        url = response.url
        if any(k in url for k in [
            "rate", "fee", "quote", "transfer", "price", "currency", "euronet"
        ]):
            try:
                body = response.json()
                if isinstance(body, dict) and len(body) > 1:
                    results.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    try:
        url = (
            f"https://www.euronet.eu/send-money"
            f"?from={from_ccy}&to={to_ccy}&amount={int(amount)}"
        )
        page.goto(url, wait_until="networkidle", timeout=40000)
        time.sleep(4)
        page.close()

        for body in results:
            rate = (body.get("rate") or body.get("exchangeRate") or
                    body.get("fxRate"))
            fee = (body.get("fee") or body.get("transferFee") or 0)
            received = (body.get("receivedAmount") or body.get("destinationAmount") or
                        body.get("targetAmount"))
            if rate:
                q = Quote(
                    "Euronet", from_ccy, to_ccy, amount,
                    fee_usd=float(fee or 0),
                    fx_rate=float(rate),
                    received_amount=float(received) if received else None,
                    note="Bank transfer",
                )
                q.fill_gaps()
                return q

        return Quote("Euronet", from_ccy, to_ccy, amount,
                     error="Corridor not supported or page unavailable")

    except Exception as e:
        try: page.close()
        except Exception: pass
        return Quote("Euronet", from_ccy, to_ccy, amount, error=str(e)[:120])


# ── Orchestrator ───────────────────────────────────────────────────────────────

SCRAPERS = [
    ("Wise",          scrape_wise),
    ("Remitly",       scrape_remitly),
    ("Western Union", scrape_western_union),
    ("MoneyGram",     scrape_moneygram),
    ("Revolut",       scrape_revolut),
    ("Euronet",       scrape_euronet),
]


def fetch_all_quotes(corridors, amounts):
    from playwright.sync_api import sync_playwright

    all_quotes = []

    with sync_playwright() as pw:
        browser, ctx = make_context(pw)
        try:
            for from_ccy, to_ccy in corridors:
                print(f"\n{'='*55}")
                print(f"  {from_ccy} -> {to_ccy}")
                print(f"{'='*55}")

                for name, scraper in SCRAPERS:
                    for amount in amounts:
                        print(f"  [{name}] ${amount:,} ... ", end="", flush=True)
                        try:
                            q = scraper(from_ccy, to_ccy, amount, ctx)
                            all_quotes.append(q)
                            if q.error:
                                print(f"ERROR: {q.error[:70]}")
                            else:
                                rcv = (f"{q.received_amount:,.2f} {to_ccy}"
                                       if q.received_amount else "?")
                                fee = (f"${q.fee_usd:.2f}"
                                       if q.fee_usd is not None else "?")
                                mkp = (f"{q.fx_markup_pct:.2f}%"
                                       if q.fx_markup_pct is not None else "?")
                                print(f"OK  fee={fee}  mkp={mkp}  rcv={rcv}")
                        except Exception as e:
                            q = Quote(name, from_ccy, to_ccy, amount,
                                      error=str(e)[:80])
                            all_quotes.append(q)
                            print(f"ERROR: {str(e)[:70]}")
                        time.sleep(2)
        finally:
            ctx.close()
            browser.close()

    return all_quotes


# ── Excel output ───────────────────────────────────────────────────────────────

C_NAVY  = "1A3A5C"
C_BLUE  = "2E6DA4"
C_LBLUE = "D9E8F5"
C_WHITE = "FFFFFF"
C_GREEN = "E2F0D9"
C_AMBER = "FFF2CC"
C_REDBG = "FFE0E0"
C_BEST  = "C6EFCE"
C_GREY  = "F5F5F5"
C_DKGRY = "404040"


def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)


def _font(bold=False, color=C_DKGRY, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")


def _border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)


def write_excel(all_quotes, amounts, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Summary"
    _write_summary(ws, all_quotes, amounts)

    corridors_seen = list(dict.fromkeys(
        f"{q.from_ccy}->{q.to_ccy}" for q in all_quotes
    ))
    for cor in corridors_seen:
        ws2 = wb.create_sheet(title=cor)
        _write_raw(ws2, [q for q in all_quotes
                         if f"{q.from_ccy}->{q.to_ccy}" == cor])

    wl = wb.create_sheet("Legend")
    _write_legend(wl)
    wb.save(filename)
    print(f"\nSaved: {filename}")


def _write_summary(ws, all_quotes, amounts):
    by_corridor = {}
    for q in all_quotes:
        cor = f"{q.from_ccy}->{q.to_ccy}"
        by_corridor.setdefault(cor, {}).setdefault(
            q.provider, {})[q.send_amount] = q

    CPAMT = 4
    total_cols = 1 + len(amounts) * CPAMT
    row = 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    ws.cell(row, 1).value = "FT Partners - Remittance Competitive Pricing Tracker"
    ws.cell(row, 1).font = Font(bold=True, size=14,
                                color=C_WHITE, name="Arial")
    ws.cell(row, 1).fill = _fill(C_NAVY)
    ws.cell(row, 1).alignment = Alignment(horizontal="center",
                                          vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    ws.cell(row, 1).value = (
        f"Bank-to-bank transfers  |  Prices scraped live from each "
        f"provider website  |  Generated: "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws.cell(row, 1).font = _font(italic=True, color=C_WHITE, size=9)
    ws.cell(row, 1).fill = _fill(C_BLUE)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    row += 2

    for corridor, providers in by_corridor.items():
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)
        ws.cell(row, 1).value = f"  {corridor}  |  Bank Account Delivery"
        ws.cell(row, 1).font = Font(bold=True, size=11,
                                    color=C_WHITE, name="Arial")
        ws.cell(row, 1).fill = _fill(C_BLUE)
        ws.cell(row, 1).alignment = Alignment(horizontal="left",
                                              vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        ws.cell(row, 1).value = "Provider"
        ws.cell(row, 1).font = _font(bold=True, color=C_WHITE)
        ws.cell(row, 1).fill = _fill(C_NAVY)
        ws.cell(row, 1).alignment = Alignment(horizontal="center",
                                              vertical="center")
        ws.cell(row, 1).border = _border()
        col = 2
        for amt in amounts:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + CPAMT - 1)
            c = ws.cell(row, col)
            c.value = f"${amt:,}"
            c.font = _font(bold=True, color=C_WHITE)
            c.fill = _fill(C_NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            col += CPAMT
        ws.row_dimensions[row].height = 18
        row += 1

        ws.cell(row, 1).fill = _fill(C_LBLUE)
        ws.cell(row, 1).border = _border()
        col = 2
        for _ in amounts:
            for lbl in ["Fee ($)", "FX Rate", "Mkp %", "Received"]:
                c = ws.cell(row, col)
                c.value = lbl
                c.font = _font(bold=True, size=8)
                c.fill = _fill(C_LBLUE)
                c.alignment = Alignment(horizontal="center")
                c.border = _border()
                col += 1
        ws.row_dimensions[row].height = 14
        row += 1

        best_rcv, best_fee = {}, {}
        for pname, pdata in providers.items():
            for amt in amounts:
                q = pdata.get(amt)
                if not q or q.error:
                    continue
                if (q.received_amount and
                        (amt not in best_rcv or
                         q.received_amount > best_rcv[amt])):
                    best_rcv[amt] = q.received_amount
                if (q.fee_usd is not None and
                        (amt not in best_fee or
                         q.fee_usd < best_fee[amt])):
                    best_fee[amt] = q.fee_usd

        for i, pname in enumerate(sorted(providers.keys())):
            pdata = providers[pname]
            bg = C_WHITE if i % 2 == 0 else C_GREY
            ws.cell(row, 1).value = pname
            ws.cell(row, 1).font = _font(bold=True)
            ws.cell(row, 1).fill = _fill(bg)
            ws.cell(row, 1).alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            ws.cell(row, 1).border = _border()

            col = 2
            for amt in amounts:
                q = pdata.get(amt)
                is_best_rcv = (q and not q.error and q.received_amount and
                               q.received_amount == best_rcv.get(amt))
                is_best_fee = (q and not q.error and
                               q.fee_usd is not None and
                               q.fee_usd == best_fee.get(amt))

                if not q or q.error:
                    err = q.error[:30] if q and q.error else "N/A"
                    for offset in range(CPAMT):
                        c = ws.cell(row, col + offset)
                        c.value = f"ERR: {err}" if offset == 0 else ""
                        c.font = _font(size=8, italic=True, color="AA0000")
                        c.fill = _fill("FFF0F0")
                        c.border = _border()
                    col += CPAMT
                    continue

                c = ws.cell(row, col)
                c.value = round(q.fee_usd, 2) if q.fee_usd is not None else 0
                c.number_format = "$#,##0.00"
                c.font = _font(bold=is_best_fee,
                               color="006600" if is_best_fee else C_DKGRY)
                c.fill = _fill(C_BEST if is_best_fee else bg)
                c.alignment = Alignment(horizontal="right")
                c.border = _border()
                col += 1

                c = ws.cell(row, col)
                c.value = round(q.fx_rate, 4) if q.fx_rate else 0
                c.number_format = "0.0000"
                c.font = _font()
                c.fill = _fill(bg)
                c.alignment = Alignment(horizontal="right")
                c.border = _border()
                col += 1

                c = ws.cell(row, col)
                mkp = q.fx_markup_pct
                c.value = round(mkp / 100, 4) if mkp is not None else 0
                c.number_format = "0.00%"
                if mkp is not None:
                    if mkp < 0.5:
                        c.fill = _fill(C_GREEN)
                        c.font = _font(color="276221", bold=True)
                    elif mkp < 1.5:
                        c.fill = _fill(C_AMBER)
                        c.font = _font(color="7D6608", bold=True)
                    else:
                        c.fill = _fill(C_REDBG)
                        c.font = _font(color="B22222", bold=True)
                else:
                    c.fill = _fill(bg)
                    c.font = _font()
                c.alignment = Alignment(horizontal="right")
                c.border = _border()
                col += 1

                c = ws.cell(row, col)
                c.value = (round(q.received_amount, 2)
                           if q.received_amount else 0)
                c.number_format = "#,##0.00"
                c.font = _font(bold=is_best_rcv,
                               color="006600" if is_best_rcv else C_DKGRY)
                c.fill = _fill(C_BEST if is_best_rcv else bg)
                c.alignment = Alignment(horizontal="right")
                c.border = _border()
                col += 1

            ws.row_dimensions[row].height = 15
            row += 1

        row += 2

    ws.column_dimensions["A"].width = 18
    ci = 2
    for _ in amounts:
        ws.column_dimensions[get_column_letter(ci)].width = 9
        ws.column_dimensions[get_column_letter(ci + 1)].width = 9
        ws.column_dimensions[get_column_letter(ci + 2)].width = 8
        ws.column_dimensions[get_column_letter(ci + 3)].width = 13
        ci += 4
    ws.freeze_panes = "B4"


def _write_raw(ws, quotes):
    headers = ["Provider", "From", "To", "Send Amount", "Fee ($)",
               "FX Rate", "FX Markup %", "Received", "Note", "Error"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col)
        c.value = h
        c.font = _font(bold=True, color=C_WHITE)
        c.fill = _fill(C_NAVY)
        c.border = _border()
        c.alignment = Alignment(horizontal="center")
    for row, q in enumerate(quotes, 2):
        vals = [
            q.provider, q.from_ccy, q.to_ccy, q.send_amount,
            q.fee_usd, q.fx_rate,
            round(q.fx_markup_pct / 100, 4)
            if q.fx_markup_pct is not None else None,
            q.received_amount, q.note, q.error,
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col)
            c.value = v
            c.font = _font()
            c.border = _border()
            if col == 4: c.number_format = "$#,##0"
            if col == 5: c.number_format = "$#,##0.00"
            if col == 6: c.number_format = "0.0000"
            if col == 7: c.number_format = "0.00%"
            if col == 8: c.number_format = "#,##0.00"
    for i, w in enumerate([18, 6, 6, 12, 10, 10, 10, 14, 45, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def _write_legend(ws):
    ws.title = "Legend"
    rows = [
        ("COLOR CODING", ""),
        ("Green FX markup", "Less than 0.5% above mid-market - excellent"),
        ("Amber FX markup", "0.5 to 1.5% above mid-market - moderate"),
        ("Red FX markup", "More than 1.5% above mid-market - expensive"),
        ("Green background (Received)",
         "Best received amount for that send amount"),
        ("Green background (Fee)", "Lowest fee for that send amount"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("Method",
         "All prices scraped live from each provider public website "
         "using headless Chrome browser"),
        ("Wise", "wise.com - mid-market rate, transparent fee"),
        ("Remitly", "remitly.com - Economy rate, bank deposit"),
        ("Western Union",
         "westernunion.com - bank account in and out, online pricing"),
        ("MoneyGram", "moneygram.com - bank deposit delivery"),
        ("Revolut",
         "revolut.com - Standard free plan, external bank transfer"),
        ("Euronet", "euronet.eu - coverage varies by corridor"),
        ("", ""),
        ("NOTES", ""),
        ("FX Markup",
         "Percent above mid-market rate. Lower is better for recipient."),
        ("Received",
         "Amount recipient gets after all fees and FX markup applied."),
        ("Bank Transfer",
         "All pricing is bank account send plus bank account receive. "
         "Not cash pickup."),
        ("Revolut",
         "Standard free plan shown. Premium and Metal plans have "
         "lower or no fees."),
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 85
    for i, (label, value) in enumerate(rows, 1):
        ws.cell(i, 1).value = label
        ws.cell(i, 2).value = value
        if label.isupper() and label:
            ws.cell(i, 1).font = Font(bold=True, size=10,
                                      color=C_WHITE, name="Arial")
            ws.cell(i, 1).fill = _fill(C_BLUE)
            ws.cell(i, 2).fill = _fill(C_BLUE)
        else:
            ws.cell(i, 1).font = _font(bold=bool(label))
            ws.cell(i, 2).font = _font()
        ws.row_dimensions[i].height = 15


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="FT Partners Remittance Pricing Tracker"
    )
    parser.add_argument(
        "--corridors", nargs="+", default=None,
        help="e.g. USD-EUR USD-GBP"
    )
    parser.add_argument(
        "--amounts", nargs="+", type=int, default=None,
        help="e.g. 100 500 1000"
    )
    parser.add_argument("--output", type=str, default=None)
    args = parser.parse_args()

    corridors = (
        [tuple(c.replace("-", " ").split()) for c in args.corridors]
        if args.corridors else DEFAULT_CORRIDORS
    )
    amounts = args.amounts or DEFAULT_AMOUNTS
    output_file = (
        args.output or
        f"remittance_pricing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    print("\n" + "="*55)
    print("  FT Partners - Remittance Pricing Tracker")
    print("="*55)
    print(f"  Corridors : {', '.join(f'{a}->{b}' for a, b in corridors)}")
    print(f"  Amounts   : {', '.join(f'${a:,}' for a in amounts)}")
    print(f"  Output    : {output_file}\n")

    all_quotes = fetch_all_quotes(corridors, amounts)
    errors = [q for q in all_quotes if q.error]
    print(f"\n  Total quotes : {len(all_quotes)}")
    print(f"  Errors       : {len(errors)}")
    write_excel(all_quotes, amounts, output_file)


if __name__ == "__main__":
    main()
