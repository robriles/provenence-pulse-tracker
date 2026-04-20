"""
Provenance Pulse Daily Scraper
================================
Scrapes all 12 Provenance Blockchain Metrics from https://provenance.io/pulse
on the 3m time period, and appends them with today's date to an Excel spreadsheet.

SETUP (run once):
    pip install playwright openpyxl
    playwright install chromium

SCHEDULE (run daily):
  Mac/Linux — crontab -e, add:
    0 13 * * * /usr/bin/python3 /path/to/scrape_provenance_pulse.py
  Windows — Task Scheduler, Daily at 8:00 AM ET
"""

import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── CONFIG ────────────────────────────────────────────────────────────────────
URL = "https://provenance.io/pulse"
EXCEL_PATH = Path(__file__).parent / "provenance_pulse_tracker.xlsx"
PAGE_TIMEOUT = 60_000
# ──────────────────────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "Date",
    "TVL",
    "Trading TVL",
    "3M Chain Transactions",
    "3M Chain Fees",
    "Total Participants",
    "Total Committed Value",
    "Total Loan Balance",
    "Total Loans",
    "3M Loan Amount Funded",
    "3M Loans Funded",
    "3M Loan Amount Paid",
    "3M Loans Paid",
]

# From the debug output we know the exact label text on the page.
# After clicking 3m, labels change from "Week's X" to "3 Months X".
# We match by the stable part that doesn't change between time periods.
METRIC_LABEL_PATTERNS = [
    ("TVL",                "TVL"),
    ("Trading TVL",        "Trading TVL"),
    ("3M Chain Transactions", "Chain Transactions"),
    ("3M Chain Fees",      "Chain Fees"),
    ("Total Participants", "Total Participants"),
    ("Total Committed Value", "Committed Value"),
    ("Total Loan Balance", "Loan Balance"),
    ("Total Loans",        "Total Loans"),
    ("3M Loan Amount Funded", "Loan Amount Funded"),
    ("3M Loans Funded",    "Loans Funded"),
    ("3M Loan Amount Paid","Loan Amount Paid"),
    ("3M Loans Paid",      "Loans Paid"),
]


def get_span_nodes(page):
    return page.evaluate("""
        () => {
            const spans = document.querySelectorAll('span');
            const results = [];
            for (const span of spans) {
                const text = span.textContent.trim();
                if (text.length === 0) continue;
                const rect = span.getBoundingClientRect();
                if (rect.width === 0) continue;
                results.push({ text: text, y: Math.round(rect.top), x: Math.round(rect.left) });
            }
            return results;
        }
    """)


def click_3m_and_verify(page):
    """Click the 3m button for Blockchain Metrics and verify the labels changed."""
    # There are two 3m button groups — Hash Metrics and Blockchain Metrics.
    # From debug: both sets are at y~582 and y~1352. We want the second (higher y).
    for attempt in range(3):
        try:
            buttons = page.locator("button").filter(has_text=re.compile(r"^3m$")).all()
            print(f"  Found {len(buttons)} '3m' buttons (attempt {attempt+1})")
            if len(buttons) >= 2:
                # Click the second 3m button (Blockchain Metrics section)
                buttons[1].click()
            elif len(buttons) == 1:
                buttons[0].click()
            else:
                print("  No 3m buttons found!")
                return False

            # Wait for labels to update — look for "3 Months" text appearing
            page.wait_for_timeout(5000)

            # Verify by checking if any span now contains "3 Months"
            nodes = get_span_nodes(page)
            has_3m = any("3 Months" in n['text'] or "months" in n['text'].lower() for n in nodes)
            if has_3m:
                print("  ✅ 3m view confirmed active")
                return True
            else:
                print(f"  ⚠️  3m labels not detected yet, retrying...")
                page.wait_for_timeout(3000)

        except Exception as e:
            print(f"  Button click error: {e}")
            page.wait_for_timeout(2000)

    print("  Proceeding despite unconfirmed 3m state")
    return False


def extract_metrics_from_nodes(nodes):
    """Match label patterns to values using position-aware span extraction."""
    results = {h: "N/A" for h in EXCEL_HEADERS[1:]}

    # Only look at spans in the Blockchain Metrics section (y > 1200 from debug)
    # This avoids accidentally matching Hash Metrics labels like "Total Supply"
    blockchain_nodes = [n for n in nodes if n['y'] > 1200]
    print(f"  Blockchain section spans: {len(blockchain_nodes)}")

    for header, pattern in METRIC_LABEL_PATTERNS:
        found = False
        for j, node in enumerate(blockchain_nodes):
            node_text = node['text']
            if pattern.lower() in node_text.lower() and len(node_text) < 80:
                # Value span is typically the next numeric span at a lower y position
                label_y = node['y']
                label_x = node['x']

                # Look for the value: a span below this label (higher y),
                # similar x position, containing a number
                for k in range(j + 1, min(j + 8, len(blockchain_nodes))):
                    candidate = blockchain_nodes[k]
                    c_text = candidate['text'].strip()

                    # Skip 'i' (info icon), pure percentages, and change indicators
                    if c_text == 'i':
                        continue
                    if re.match(r'^\(.*%\)$', c_text):
                        continue
                    if re.match(r'^[\+\-]?\d+(\.\d+)?%$', c_text):
                        continue

                    # Must contain digits
                    if not re.search(r'\d', c_text):
                        continue

                    # Must look like a real value (dollar, number with commas, etc.)
                    if re.search(r'[\$\d,\.]+[BMKTbmkt]?', c_text):
                        results[header] = c_text
                        print(f"  {header}: {c_text}")
                        found = True
                        break

                if found:
                    break

        if not found:
            print(f"  {header}: NOT FOUND")

    return results


def scrape_metrics() -> dict:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        print("Loading page...")
        try:
            page.goto(URL, timeout=PAGE_TIMEOUT, wait_until="networkidle")
        except PlaywrightTimeout:
            page.goto(URL, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
            page.wait_for_timeout(10000)

        print("Clicking 3m button...")
        click_3m_and_verify(page)

        print("Extracting metrics...")
        nodes = get_span_nodes(page)
        print(f"Total spans on page: {len(nodes)}")

        results = extract_metrics_from_nodes(nodes)
        browser.close()

    return results


def get_or_create_workbook():
    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pulse Metrics"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, h in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 36

    wb.save(EXCEL_PATH)
    return wb


def append_row(wb, today: date, metrics: dict):
    ws = wb.active
    next_row = ws.max_row + 1
    row_data = [today.isoformat()] + [metrics[h] for h in EXCEL_HEADERS[1:]]

    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(horizontal="center")

    wb.save(EXCEL_PATH)
    print(f"\n✅ Saved {len(metrics)} metrics for {today.isoformat()}")


def main():
    today = date.today()
    print(f"Scraping {URL} for {today} (3m view)...")

    wb = get_or_create_workbook()
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == today.isoformat():
            print(f"⚠️  Entry for {today} already exists. Skipping.")
            return

    try:
        metrics = scrape_metrics()
    except Exception as e:
        print(f"❌ Fatal scrape error: {e}", file=sys.stderr)
        metrics = {h: "SCRAPE FAILED" for h in EXCEL_HEADERS[1:]}

    append_row(wb, today, metrics)


if __name__ == "__main__":
    main()
